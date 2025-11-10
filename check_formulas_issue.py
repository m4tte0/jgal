import openpyxl
from pathlib import Path

def check_formulas_in_output():
    """Check for formula issues in the output file"""
    output_file = Path("Avanzamento_schede_automated.xlsx")

    print(f"Checking formulas in: {output_file}")

    wb = openpyxl.load_workbook(output_file, data_only=False)  # Keep formulas
    ws = wb.active

    # Check the Delta column (should be column 7 now)
    print("\nChecking Delta column (Column 7):")
    for row_idx in range(1, min(11, ws.max_row + 1)):
        cell = ws.cell(row_idx, 7)
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"  Row {row_idx}: {cell.value}")

    # Check all columns for formulas
    print("\nAll formulas in first 10 rows:")
    for row_idx in range(1, 11):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"  Row {row_idx}, Col {col_idx}: {cell.value}")

    wb.close()

check_formulas_in_output()
