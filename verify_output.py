import openpyxl
from pathlib import Path

def verify_output():
    """Verify the generated output file"""
    output_file = Path("Avanzamento_schede_automated.xlsx")

    print(f"Verifying output file: {output_file}")

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Show header row
    print("\nHeader row:")
    for col_idx in range(1, min(ws.max_column + 1, 12)):  # Show first 11 columns
        cell = ws.cell(1, col_idx)
        print(f"  Col {col_idx}: {cell.value}")

    # Show a few data rows with Matricola and first few date columns
    print("\nSample data rows (Matricola + first 3 date columns):")
    matricola_col = 6  # Column F
    date_start_col = 9  # Column I (first date column)

    for row_idx in range(2, min(7, ws.max_row + 1)):  # Show first 5 data rows
        matricola = ws.cell(row_idx, matricola_col).value
        dates = []
        for col_idx in range(date_start_col, date_start_col + 3):
            date_val = ws.cell(row_idx, col_idx).value
            dates.append(str(date_val) if date_val else "None")

        print(f"  Row {row_idx}: Matricola={matricola}, Dates=[{', '.join(dates)}]")

    wb.close()
    print("\nVerification complete!")

verify_output()
