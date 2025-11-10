import openpyxl
from pathlib import Path

def generate_summary():
    """Generate summary report"""
    output_file = Path("Avanzamento_schede_automated.xlsx")

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    total_data_rows = ws.max_row - 1  # Excluding header
    consolidated_col = ws.max_column - 1

    filled_rows = 0
    empty_rows = 0

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, consolidated_col).value:
            filled_rows += 1
        else:
            empty_rows += 1

    print("="*60)
    print("AUTOMATION SUMMARY REPORT")
    print("="*60)
    print(f"\nGenerated file: {output_file}")
    print(f"\nTotal columns: {ws.max_column}")
    print(f"  - Columns A-H: Original data (excluding removed date columns)")
    print(f"  - Columns I-X: 16 Planning date columns (2025-07-04 to 2025-10-31)")
    print(f"  - Column Y: Consolidated 'Data prevista avanzamento'")
    print(f"  - Column Z: 'Data effettiva avanzamento' (empty)")

    print(f"\nTotal data rows: {total_data_rows}")
    print(f"  - Rows with consolidated date: {filled_rows} ({filled_rows/total_data_rows*100:.1f}%)")
    print(f"  - Rows without consolidated date: {empty_rows} ({empty_rows/total_data_rows*100:.1f}%)")

    print(f"\nMatching strategy:")
    print(f"  1. Match by Matricola (primary)")
    print(f"  2. Match by Articolo (fallback)")
    print(f"  3. Ignore 'KOM' values")
    print(f"  4. Use last valid Planning date (2025-10-31 backwards)")

    print(f"\n{'='*60}")
    wb.close()

generate_summary()
