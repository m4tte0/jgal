import openpyxl
from pathlib import Path

def find_column_with_label(file_path, search_terms):
    """Find column containing specific text in headers"""
    print(f"\nSearching in: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Search in first 10 rows for headers
    matches = []
    for row_idx in range(1, min(11, ws.max_row + 1)):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            if cell.value:
                cell_text = str(cell.value)
                for term in search_terms:
                    if term.lower() in cell_text.lower():
                        matches.append({
                            'row': row_idx,
                            'col': col_idx,
                            'value': cell_text,
                            'term': term
                        })

    if matches:
        print(f"Found {len(matches)} matches:")
        for m in matches:
            print(f"  Row {m['row']}, Col {m['col']}: '{m['value']}' (matched '{m['term']}')")
            # Show surrounding context
            print(f"    Context: ", end="")
            for offset in [-1, 0, 1]:
                if m['col'] + offset > 0:
                    context_cell = ws.cell(m['row'], m['col'] + offset)
                    print(f"[{context_cell.value}] ", end="")
            print()
    else:
        print("No matches found")

    wb.close()
    return matches

# Search terms
search_terms = ["Rilascio", "DiBa", "Disegni", "Mecc", "Idr"]

planning_file = Path("_ref/usbilli/Planning/Planning_25_07_04.xlsx")
find_column_with_label(planning_file, search_terms)

# Let's also search for "Matricola" to confirm its location
print("\n" + "="*80)
print("Confirming 'Matricola' location:")
find_column_with_label(planning_file, ["Matricola"])
