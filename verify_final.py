import openpyxl
from pathlib import Path

def verify_final():
    """Verify all rows are filled"""
    output_file = Path("Avanzamento_schede_automated.xlsx")

    print(f"Verifying final output: {output_file}")

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Check consolidated column (column Y, which is max_column - 1)
    consolidated_col = ws.max_column - 1
    articolo_col = 1
    matricola_col = 6

    print(f"\nChecking consolidated column (Column {consolidated_col}):")

    rows_with_data = 0
    rows_without_data = 0
    rows_without_data_list = []

    for row_idx in range(2, ws.max_row + 1):
        consolidated_val = ws.cell(row_idx, consolidated_col).value
        articolo = ws.cell(row_idx, articolo_col).value
        matricola = ws.cell(row_idx, matricola_col).value

        if consolidated_val:
            rows_with_data += 1
        else:
            rows_without_data += 1
            if rows_without_data <= 10:  # Show first 10 empty rows
                rows_without_data_list.append({
                    'row': row_idx,
                    'articolo': articolo,
                    'matricola': matricola
                })

    print(f"  Rows WITH consolidated date: {rows_with_data}")
    print(f"  Rows WITHOUT consolidated date: {rows_without_data}")

    if rows_without_data > 0:
        print(f"\n  First {min(10, rows_without_data)} rows without consolidated dates:")
        for item in rows_without_data_list:
            print(f"    Row {item['row']}: Articolo={item['articolo']}, Matricola={item['matricola']}")

            # Show what dates exist in planning columns for this row
            dates = []
            for col_idx in range(9, consolidated_col):  # All planning date columns
                date_val = ws.cell(item['row'], col_idx).value
                if date_val:
                    dates.append(str(date_val)[:10])

            if dates:
                # Check if dates are different
                unique_dates = list(set(dates))
                print(f"      Found {len(dates)} dates, {len(unique_dates)} unique: {', '.join(unique_dates[:5])}")
            else:
                print(f"      No dates found in any planning columns")

    wb.close()
    print("\nVerification complete!")

verify_final()
