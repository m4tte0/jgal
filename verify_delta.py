import openpyxl
from pathlib import Path

def verify_delta():
    """Verify Delta column calculations"""
    output_file = Path("Avanzamento_schede_automated.xlsx")

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    print("="*80)
    print("DELTA COLUMN VERIFICATION")
    print("="*80)

    # Find columns
    delta_col = None
    prevista_col = None
    effettiva_col = None

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        if header == "Delta":
            delta_col = col_idx
        elif header == "Data prevista avanzamento" and not prevista_col:
            # Get the consolidated one (second to last column)
            if col_idx == ws.max_column - 1:
                prevista_col = col_idx
        elif header == "Data effettiva avanzamento":
            effettiva_col = col_idx

    print(f"\nColumn indices:")
    print(f"  Delta: {delta_col}")
    print(f"  Data prevista (consolidated): {prevista_col}")
    print(f"  Data effettiva: {effettiva_col}")

    # Verify calculations for first 10 rows
    print(f"\nSample delta calculations (first 10 rows with all values):")
    print(f"{'Row':<5} {'Prevista':<12} {'Effettiva':<12} {'Delta':<8} {'Verification':<15}")
    print("-" * 80)

    shown = 0
    for row_idx in range(2, ws.max_row + 1):
        prevista = ws.cell(row_idx, prevista_col).value
        effettiva = ws.cell(row_idx, effettiva_col).value
        delta = ws.cell(row_idx, delta_col).value

        if prevista and effettiva and shown < 10:
            # Calculate expected delta
            expected_delta = (effettiva - prevista).days if hasattr(effettiva, 'date') and hasattr(prevista, 'date') else None

            prevista_str = str(prevista)[:10]
            effettiva_str = str(effettiva)[:10]
            delta_str = str(delta) if delta is not None else "None"

            # Verify
            verification = "OK" if delta == expected_delta else f"ERROR (expected {expected_delta})"

            print(f"{row_idx:<5} {prevista_str:<12} {effettiva_str:<12} {delta_str:<8} {verification:<15}")
            shown += 1

    # Statistics
    delta_filled = 0
    delta_empty = 0
    positive_deltas = 0
    negative_deltas = 0
    zero_deltas = 0

    for row_idx in range(2, ws.max_row + 1):
        delta = ws.cell(row_idx, delta_col).value
        if delta is not None:
            delta_filled += 1
            if delta > 0:
                positive_deltas += 1
            elif delta < 0:
                negative_deltas += 1
            else:
                zero_deltas += 1
        else:
            delta_empty += 1

    print(f"\n{'='*80}")
    print(f"DELTA STATISTICS")
    print(f"{'='*80}")
    print(f"Total rows: {ws.max_row - 1}")
    print(f"Delta filled: {delta_filled} ({delta_filled/(ws.max_row-1)*100:.1f}%)")
    print(f"Delta empty: {delta_empty} ({delta_empty/(ws.max_row-1)*100:.1f}%)")
    print(f"\nDelta distribution:")
    print(f"  Positive (late): {positive_deltas}")
    print(f"  Zero (on time): {zero_deltas}")
    print(f"  Negative (early): {negative_deltas}")
    print(f"{'='*80}")

    wb.close()

verify_delta()
