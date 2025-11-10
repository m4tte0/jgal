import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from pathlib import Path

def check_formulas(file_path):
    """Check what the formulas in col 31 reference"""
    print(f"\nChecking formulas in: {file_path}")

    wb = openpyxl.load_workbook(file_path, data_only=False)  # Keep formulas
    ws = wb.active

    # Column 31 has formulas pointing to other columns
    # Let's see what AF (column 32 in Excel, which is column index 32) contains
    print("\nColumn AF (index 32) - should be referenced by col 31:")
    for row_idx in range(1, 11):
        cell = ws.cell(row_idx, 32)
        print(f"Row {row_idx}: {cell.value}")

    # Let's find out what column letter is column 31
    print(f"\nColumn 31 is letter: {get_column_letter(31)}")
    print(f"Column AF is index: {column_index_from_string('AF')}")

    # Now load with data_only=True to get calculated values
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    ws_data = wb_data.active

    print("\nColumn 31 with calculated values (data_only=True):")
    for row_idx in range(1, 11):
        cell = ws_data.cell(row_idx, 31)
        print(f"Row {row_idx}: {cell.value}")

    wb.close()
    wb_data.close()

planning_file = Path("_ref/usbilli/Planning/Planning_25_07_04.xlsx")
check_formulas(planning_file)
