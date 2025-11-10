import openpyxl
from pathlib import Path

def inspect_excel(file_path, max_rows=20):
    """Inspect Excel file structure"""
    print(f"\n{'='*80}")
    print(f"Inspecting: {file_path}")
    print('='*80)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    print(f"Sheet name: {ws.title}")
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")

    # Print headers (first few rows)
    print(f"\nFirst {min(max_rows, ws.max_row)} rows:")
    for row_idx in range(1, min(max_rows + 1, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(ws.max_column + 1, 15)):  # Limit to first 15 columns
            cell = ws.cell(row_idx, col_idx)
            value = cell.value
            if value is None:
                value = ""
            row_data.append(str(value)[:20])  # Limit to 20 chars per cell
        print(f"Row {row_idx:2d}: {' | '.join(row_data)}")

    wb.close()

# Inspect source file
source_file = Path("_ref/usbilli/Avanzamento schede 3° trimestre 2025.xlsx")
inspect_excel(source_file, max_rows=10)

# Inspect one Planning file
planning_file = Path("_ref/usbilli/Planning/Planning_25_07_04.xlsx")
inspect_excel(planning_file, max_rows=10)

# Inspect generated file
output_file = Path("Avanzamento_schede_automated.xlsx")
if output_file.exists():
    inspect_excel(output_file, max_rows=10)
