import openpyxl
from pathlib import Path

def verify_consolidated():
    """Verify the consolidated column"""
    output_file = Path("Avanzamento_schede_automated.xlsx")

    print(f"Verifying consolidated column in: {output_file}")

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Show last few column headers
    print("\nLast 5 column headers:")
    for col_idx in range(ws.max_column - 4, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        print(f"  Column {col_idx}: {header}")

    # Show sample data for rows with consolidated dates
    print("\nSample rows with consolidated dates (showing Matricola, last Planning date, Consolidated, Effettiva):")
    matricola_col = 6
    last_planning_col = ws.max_column - 2  # Column X (last planning date)
    consolidated_col = ws.max_column - 1   # Column Y (consolidated)
    effettiva_col = ws.max_column          # Column Z (effettiva)

    rows_shown = 0
    for row_idx in range(2, ws.max_row + 1):
        consolidated_val = ws.cell(row_idx, consolidated_col).value
        if consolidated_val and rows_shown < 10:
            matricola = ws.cell(row_idx, matricola_col).value
            last_planning = ws.cell(row_idx, last_planning_col).value
            effettiva = ws.cell(row_idx, effettiva_col).value

            print(f"  Row {row_idx}: Matricola={matricola}, Last Planning={last_planning}, Consolidated={consolidated_val}, Effettiva={effettiva}")
            rows_shown += 1

    # Show example of row with multiple different dates (no consolidated date)
    print("\nSample rows WITHOUT consolidated dates (multiple different dates found):")
    rows_shown = 0
    for row_idx in range(2, ws.max_row + 1):
        consolidated_val = ws.cell(row_idx, consolidated_col).value
        if not consolidated_val and rows_shown < 5:
            matricola = ws.cell(row_idx, matricola_col).value
            if matricola:
                # Show a few planning dates for this row
                dates = []
                for col_idx in range(9, min(13, ws.max_column)):  # Show first 4 planning dates
                    date_val = ws.cell(row_idx, col_idx).value
                    dates.append(str(date_val) if date_val else "None")

                print(f"  Row {row_idx}: Matricola={matricola}, First 4 dates=[{', '.join(dates)}]")
                rows_shown += 1

    wb.close()
    print("\nVerification complete!")

verify_consolidated()
