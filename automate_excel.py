import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.utils import get_column_letter
import re
import os
from pathlib import Path
from copy import copy
import csv
from datetime import datetime

def find_jgal_file(jgal_folder, articolo, revisione):
    """
    Find the matching CSV file in jgal folder based on Articolo and Revisione.

    Matching logic:
    1. Replace "/" with "_" in articolo name (filesystem compatibility)
    2. If Revisione > 0: First try {Articolo}_rev{Revisione}.csv
    3. If not found or Revisione = 0: Try {Articolo}.csv
    4. Return the matched file path or None if not found
    """
    # Convert to string and replace "/" with "_" for filesystem compatibility
    articolo_normalized = str(articolo).replace('/', '_')

    # Always try with revision suffix first (including _rev0)
    if revisione is not None:
        rev_file = jgal_folder / f"{articolo_normalized}_rev{int(revisione)}.csv"
        if rev_file.exists():
            return rev_file

    # Try without revision suffix as fallback
    base_file = jgal_folder / f"{articolo_normalized}.csv"
    if base_file.exists():
        return base_file

    return None

def extract_date_from_jgal_csv(csv_file):
    """
    Extract the date from a jgal CSV file where Sequenza = 90.

    Returns a datetime object or None if not found.
    """
    try:
        with open(csv_file, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.DictReader(f, delimiter=';')

            for row in reader:
                sequenza = row.get('Sequenza', '').strip()
                if sequenza == '90':
                    date_str = row.get('Data', '').strip()
                    if date_str:
                        # Parse date in format DD/MM/YY
                        try:
                            date_obj = datetime.strptime(date_str, '%d/%m/%y')
                            return date_obj
                        except ValueError:
                            # Try other common formats if needed
                            pass
                    break
    except Exception as e:
        print(f"Error reading {csv_file}: {e}")

    return None

def extract_date_from_filename(filename):
    """Extract date from Planning filename in format Planning_yy_mm_dd.xlsx"""
    match = re.search(r'Planning_(\d{2})_(\d{2})_(\d{2})\.xlsx', filename)
    if match:
        yy, mm, dd = match.groups()
        return f"20{yy}-{mm}-{dd}"
    return None

def copy_cell_style(source_cell, target_cell):
    """Copy all style attributes from source to target cell"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def main():
    # Define paths
    base_path = Path("_ref/usbilli")
    source_file = base_path / "Avanzamento schede 3° trimestre 2025.xlsx"
    planning_folder = base_path / "Planning"
    output_file = "Avanzamento_schede_automated.xlsx"

    # Get all Planning files and extract dates
    planning_files = sorted(planning_folder.glob("Planning_*.xlsx"))
    planning_dates = []
    for pf in planning_files:
        date = extract_date_from_filename(pf.name)
        if date:
            planning_dates.append((date, pf))

    print(f"Found {len(planning_dates)} Planning files")
    for date, pf in planning_dates:
        print(f"  - {pf.name}: {date}")

    # Load source workbook
    print(f"\nLoading source file: {source_file}")
    source_wb = openpyxl.load_workbook(source_file)
    source_ws = source_wb.active

    # Find columns to exclude
    columns_to_exclude = []
    header_row = 1  # Assuming headers are in row 1

    for col_idx in range(1, source_ws.max_column + 1):
        cell_value = source_ws.cell(header_row, col_idx).value
        if cell_value and ("Data effettiva avanzamento" in str(cell_value) or
                          "Data prevista avanzamento" in str(cell_value)):
            columns_to_exclude.append(col_idx)
            print(f"Column to exclude: {get_column_letter(col_idx)} - {cell_value}")

    # Create new workbook
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = source_ws.title

    # Copy all data except excluded columns
    print("\nCopying data and formatting...")
    col_mapping = {}  # Maps old column index to new column index
    new_col_idx = 1

    for old_col_idx in range(1, source_ws.max_column + 1):
        if old_col_idx in columns_to_exclude:
            continue

        col_mapping[old_col_idx] = new_col_idx

        # Copy column width
        old_col_letter = get_column_letter(old_col_idx)
        new_col_letter = get_column_letter(new_col_idx)
        if old_col_letter in source_ws.column_dimensions:
            new_ws.column_dimensions[new_col_letter].width = source_ws.column_dimensions[old_col_letter].width

        # Copy all cells in this column
        for row_idx in range(1, source_ws.max_row + 1):
            source_cell = source_ws.cell(row_idx, old_col_idx)
            target_cell = new_ws.cell(row_idx, new_col_idx)

            # Copy value (but skip formulas that reference excluded columns)
            cell_value = source_cell.value
            if cell_value and isinstance(cell_value, str) and cell_value.startswith('='):
                # This is a formula - check if it references excluded columns
                skip_formula = False
                for excluded_col in columns_to_exclude:
                    excluded_letter = get_column_letter(excluded_col)
                    if excluded_letter in cell_value:
                        skip_formula = True
                        break

                if skip_formula:
                    # Clear the formula to avoid circular references
                    target_cell.value = None
                else:
                    target_cell.value = cell_value
            else:
                target_cell.value = cell_value

            # Copy style
            copy_cell_style(source_cell, target_cell)

        new_col_idx += 1

    # Copy row heights
    for row_idx in range(1, source_ws.max_row + 1):
        if row_idx in source_ws.row_dimensions:
            new_ws.row_dimensions[row_idx].height = source_ws.row_dimensions[row_idx].height

    # Add "Data prevista avanzamento" column for each Planning file
    print(f"\nAdding {len(planning_dates)} 'Data prevista avanzamento' columns...")

    # First, find the Matricola and Articolo columns in the new worksheet
    matricola_col_idx = None
    articolo_col_idx = None
    for col_idx in range(1, new_ws.max_column + 1):
        header_value = new_ws.cell(header_row, col_idx).value
        if header_value == "Matricola":
            matricola_col_idx = col_idx
        elif header_value == "Articolo":
            articolo_col_idx = col_idx

    if not matricola_col_idx or not articolo_col_idx:
        print("ERROR: Could not find 'Matricola' or 'Articolo' column in source file!")
        return None

    print(f"Found 'Matricola' column at index {matricola_col_idx}")
    print(f"Found 'Articolo' column at index {articolo_col_idx}")

    for idx, (date, planning_file) in enumerate(planning_dates):
        col_idx = new_col_idx + idx
        col_letter = get_column_letter(col_idx)

        # Set header
        header_cell = new_ws.cell(header_row, col_idx)
        if len(planning_dates) == 1:
            header_cell.value = "Data prevista avanzamento"
        else:
            header_cell.value = f"Data prevista avanzamento ({date})"

        # Copy header style from first column
        first_header = new_ws.cell(header_row, 1)
        copy_cell_style(first_header, header_cell)

        # Set column width
        new_ws.column_dimensions[col_letter].width = 20

        print(f"  Processing column {col_letter}: {header_cell.value}")

        # Load Planning file and extract dates
        print(f"    Loading Planning file: {planning_file}")
        planning_wb = openpyxl.load_workbook(planning_file, data_only=True)
        planning_ws = planning_wb.active

        # Build mappings from Planning file
        # Column 2 = Matricola, Column 4 = Articolo, Column 31 = Rilascio DiBa/Disegni (Mecc. + Idr.)
        matricola_to_date = {}
        articolo_to_date = {}
        planning_data_start_row = 5  # Data starts at row 5

        for row_idx in range(planning_data_start_row, planning_ws.max_row + 1):
            matricola_cell = planning_ws.cell(row_idx, 2)  # Column 2 = Matricola
            articolo_cell = planning_ws.cell(row_idx, 4)   # Column 4 = Articolo
            date_cell = planning_ws.cell(row_idx, 31)      # Column 31 = Rilascio DiBa/Disegni

            if matricola_cell.value:
                matricola = str(matricola_cell.value).strip()
                date_value = date_cell.value
                matricola_to_date[matricola] = date_value

            if articolo_cell.value:
                articolo = str(articolo_cell.value).strip()
                date_value = date_cell.value
                articolo_to_date[articolo] = date_value

        print(f"    Found {len(matricola_to_date)} matricola and {len(articolo_to_date)} articolo entries in Planning file")

        # Now populate the new worksheet by matching Matricola first, then Articolo as fallback
        matches_by_matricola = 0
        matches_by_articolo = 0
        for row_idx in range(2, new_ws.max_row + 1):  # Start from row 2 (skip header)
            target_matricola_cell = new_ws.cell(row_idx, matricola_col_idx)
            target_articolo_cell = new_ws.cell(row_idx, articolo_col_idx)
            target_date_cell = new_ws.cell(row_idx, col_idx)

            date_value = None

            # Try matching by Matricola first
            if target_matricola_cell.value:
                target_matricola = str(target_matricola_cell.value).strip()
                if target_matricola in matricola_to_date:
                    date_value = matricola_to_date[target_matricola]
                    matches_by_matricola += 1

            # If no match by Matricola, try Articolo
            if not date_value and target_articolo_cell.value:
                target_articolo = str(target_articolo_cell.value).strip()
                if target_articolo in articolo_to_date:
                    date_value = articolo_to_date[target_articolo]
                    matches_by_articolo += 1

            # Populate the cell if we found a match
            if date_value:
                target_date_cell.value = date_value
                # Copy number format for dates
                if date_value:
                    target_date_cell.number_format = 'YYYY-MM-DD'

        print(f"    Matched {matches_by_matricola} rows by Matricola, {matches_by_articolo} rows by Articolo")

        planning_wb.close()

    # Add consolidated "Data prevista avanzamento" column (no date in label)
    consolidated_col_idx = new_col_idx + len(planning_dates)
    consolidated_col_letter = get_column_letter(consolidated_col_idx)

    consolidated_header = new_ws.cell(header_row, consolidated_col_idx)
    consolidated_header.value = "Data prevista avanzamento"

    first_header = new_ws.cell(header_row, 1)
    copy_cell_style(first_header, consolidated_header)
    new_ws.column_dimensions[consolidated_col_letter].width = 20

    print(f"\nAdding consolidated column {consolidated_col_letter}: Data prevista avanzamento")

    # Populate consolidated column using the last Planning file date, ignoring 'KOM' values
    consolidated_count = 0
    for row_idx in range(2, new_ws.max_row + 1):  # Start from row 2 (skip header)
        # Collect valid dates from the planning columns (ignore 'KOM' and non-date values)
        # Start from the last planning file and work backwards
        last_valid_date = None

        for col_offset in range(len(planning_dates) - 1, -1, -1):  # Iterate backwards from last to first
            date_col_idx = new_col_idx + col_offset
            date_cell = new_ws.cell(row_idx, date_col_idx)
            cell_value = date_cell.value

            # Check if it's a valid date (not 'KOM' and not None)
            if cell_value and str(cell_value).strip().upper() != 'KOM':
                # Check if it's a datetime object (actual date)
                if hasattr(cell_value, 'year'):  # datetime object
                    last_valid_date = cell_value
                    break

        # Populate consolidated column if we found a valid date
        if last_valid_date:
            consolidated_cell = new_ws.cell(row_idx, consolidated_col_idx)
            consolidated_cell.value = last_valid_date
            consolidated_cell.number_format = 'YYYY-MM-DD'
            consolidated_count += 1

    print(f"  Populated {consolidated_count} rows with consolidated dates (using last valid Planning date)")

    # Add "Data effettiva avanzamento" column and populate from jgal CSV files
    final_col_idx = consolidated_col_idx + 1
    final_col_letter = get_column_letter(final_col_idx)

    final_header = new_ws.cell(header_row, final_col_idx)
    final_header.value = "Data effettiva avanzamento"

    first_header = new_ws.cell(header_row, 1)
    copy_cell_style(first_header, final_header)
    new_ws.column_dimensions[final_col_letter].width = 20

    print(f"\nAdding and populating column {final_col_letter}: Data effettiva avanzamento")

    # Find Articolo and Revisione columns
    articolo_col_idx = None
    revisione_col_idx = None
    for col_idx in range(1, new_ws.max_column + 1):
        header_value = new_ws.cell(header_row, col_idx).value
        if header_value == "Articolo":
            articolo_col_idx = col_idx
        elif header_value == "Revisione":
            revisione_col_idx = col_idx

    if not articolo_col_idx or not revisione_col_idx:
        print("ERROR: Could not find 'Articolo' or 'Revisione' column!")
        return None

    # Process each row to extract "Data effettiva avanzamento"
    jgal_folder = Path("_ref/jgal")
    populated_count = 0
    error_count = 0
    errors = []

    for row_idx in range(2, new_ws.max_row + 1):
        articolo = new_ws.cell(row_idx, articolo_col_idx).value
        revisione = new_ws.cell(row_idx, revisione_col_idx).value

        if not articolo:
            continue

        try:
            # Find the matching CSV file
            matching_file = find_jgal_file(jgal_folder, articolo, revisione)

            if not matching_file:
                raise Exception(f"No matching file found for Articolo={articolo}, Revisione={revisione}")

            # Extract date from CSV file (Sequenza=90)
            date_value = extract_date_from_jgal_csv(matching_file)

            if date_value:
                target_cell = new_ws.cell(row_idx, final_col_idx)
                target_cell.value = date_value
                target_cell.number_format = 'YYYY-MM-DD'
                populated_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"Row {row_idx} (Articolo={articolo}, Revisione={revisione}): {str(e)}")

    print(f"  Populated {populated_count} rows")
    if error_count > 0:
        print(f"  Errors: {error_count}")
        for err in errors[:10]:  # Show first 10 errors
            print(f"    - {err}")

    # Calculate and populate Delta column (Data effettiva - Data prevista)
    print(f"\nCalculating Delta column (Data effettiva - Data prevista)...")

    # Find Delta column
    delta_col_idx = None
    for col_idx in range(1, new_ws.max_column + 1):
        header_value = new_ws.cell(header_row, col_idx).value
        if header_value == "Delta":
            delta_col_idx = col_idx
            break

    if delta_col_idx:
        delta_populated = 0
        for row_idx in range(2, new_ws.max_row + 1):
            effettiva_val = new_ws.cell(row_idx, final_col_idx).value
            prevista_val = new_ws.cell(row_idx, consolidated_col_idx).value

            # Calculate delta if both dates exist
            if effettiva_val and prevista_val:
                try:
                    # Ensure both are datetime objects
                    if hasattr(effettiva_val, 'date') and hasattr(prevista_val, 'date'):
                        delta_days = (effettiva_val - prevista_val).days
                        delta_cell = new_ws.cell(row_idx, delta_col_idx)
                        delta_cell.value = delta_days
                        delta_cell.number_format = '0'  # Integer format
                        delta_populated += 1
                except Exception as e:
                    pass  # Skip rows with calculation errors

        print(f"  Populated {delta_populated} rows with delta values")
    else:
        print("  Warning: Delta column not found")

    # Save the new workbook
    print(f"\nSaving output file: {output_file}")
    new_wb.save(output_file)
    print("Done!")

    return output_file

if __name__ == "__main__":
    main()
