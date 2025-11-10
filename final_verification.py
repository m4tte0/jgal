import openpyxl
from pathlib import Path

def final_verification():
    """Final verification of the completed automation"""
    output_file = Path("Avanzamento_schede_automated.xlsx")

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    print("="*80)
    print("FINAL AUTOMATION VERIFICATION")
    print("="*80)

    # Column indices
    articolo_col = 1
    revisione_col = 2
    consolidated_col = ws.max_column - 1  # Column Y
    effettiva_col = ws.max_column  # Column Z

    total_rows = ws.max_row - 1  # Excluding header

    # Check consolidated column
    consolidated_filled = 0
    consolidated_empty = 0

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, consolidated_col).value:
            consolidated_filled += 1
        else:
            consolidated_empty += 1

    # Check effettiva column
    effettiva_filled = 0
    effettiva_empty = 0

    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, effettiva_col).value:
            effettiva_filled += 1
        else:
            effettiva_empty += 1

    print(f"\nTotal data rows: {total_rows}")
    print(f"\nColumn Y - 'Data prevista avanzamento' (consolidated):")
    print(f"  Filled: {consolidated_filled} ({consolidated_filled/total_rows*100:.1f}%)")
    print(f"  Empty: {consolidated_empty} ({consolidated_empty/total_rows*100:.1f}%)")

    print(f"\nColumn Z - 'Data effettiva avanzamento':")
    print(f"  Filled: {effettiva_filled} ({effettiva_filled/total_rows*100:.1f}%)")
    print(f"  Empty: {effettiva_empty} ({effettiva_empty/total_rows*100:.1f}%)")

    # Show sample data
    print(f"\nSample data (first 5 rows):")
    print(f"{'Row':<5} {'Articolo':<25} {'Rev':<5} {'Data Prev':<12} {'Data Eff':<12}")
    print("-" * 80)

    for row_idx in range(2, min(7, ws.max_row + 1)):
        articolo = str(ws.cell(row_idx, articolo_col).value)[:24]
        revisione = ws.cell(row_idx, revisione_col).value
        consolidated = ws.cell(row_idx, consolidated_col).value
        effettiva = ws.cell(row_idx, effettiva_col).value

        consolidated_str = str(consolidated)[:10] if consolidated else "None"
        effettiva_str = str(effettiva)[:10] if effettiva else "None"

        print(f"{row_idx:<5} {articolo:<25} {revisione!s:<5} {consolidated_str:<12} {effettiva_str:<12}")

    print("\n" + "="*80)
    print("AUTOMATION COMPLETE!")
    print("="*80)

    wb.close()

final_verification()
