import openpyxl
from pathlib import Path
import os

def verify_matching_logic():
    """Verify the file matching logic"""
    output_file = Path("Avanzamento_schede_automated.xlsx")
    jgal_folder = Path("_ref/jgal")

    print("Verifying file matching logic:")
    print("="*80)

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    articolo_col = 1
    revisione_col = 2

    # Check first 15 rows
    print("\nFirst 15 rows - checking file matches:")
    for row_idx in range(2, min(17, ws.max_row + 1)):
        articolo = ws.cell(row_idx, articolo_col).value
        revisione = ws.cell(row_idx, revisione_col).value

        if not articolo:
            print(f"\nRow {row_idx}: No Articolo value, SKIP")
            continue

        print(f"\nRow {row_idx}: Articolo={articolo}, Revisione={revisione}")

        # Build expected filename
        if revisione and revisione != 0 and revisione != "0":
            expected_file = f"{articolo}_rev{revisione}.csv"
        else:
            # Try without revision first
            expected_file = f"{articolo}.csv"

        expected_path = jgal_folder / expected_file

        if expected_path.exists():
            print(f"  FOUND: {expected_file}")
        else:
            # If not found, check if revision variant exists
            print(f"  NOT FOUND: {expected_file}")
            # Check for any files with this articolo
            matching_files = list(jgal_folder.glob(f"{articolo}*.csv"))
            if matching_files:
                print(f"  Available files with this Articolo: {[f.name for f in matching_files]}")

    # Now check the case with multiple revisions (ACC_2025_0475)
    print("\n" + "="*80)
    print("Checking specific case: ACC_2025_0475 with revisions")

    for row_idx in range(2, ws.max_row + 1):
        articolo = ws.cell(row_idx, articolo_col).value
        revisione = ws.cell(row_idx, revisione_col).value

        if articolo == "ACC_2025_0475":
            print(f"\nRow {row_idx}: Articolo={articolo}, Revisione={revisione}")
            if revisione == 0 or revisione == "0":
                expected_file = f"{articolo}_rev0.csv"
            else:
                expected_file = f"{articolo}_rev{revisione}.csv"

            expected_path = jgal_folder / expected_file
            print(f"  Expected file: {expected_file}")
            print(f"  Exists: {expected_path.exists()}")

    wb.close()

verify_matching_logic()
