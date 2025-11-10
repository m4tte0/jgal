import openpyxl
from pathlib import Path

def check_articolo_in_planning():
    """Check if Articolo column exists in Planning files"""
    planning_file = Path("_ref/usbilli/Planning/Planning_25_07_04.xlsx")

    print(f"Checking for Articolo column in: {planning_file}")

    wb = openpyxl.load_workbook(planning_file)
    ws = wb.active

    # Search for Articolo in first few rows
    print("\nFirst 10 columns of first 5 rows:")
    for row_idx in range(1, 6):
        print(f"Row {row_idx}:", end=" ")
        for col_idx in range(1, 11):
            cell = ws.cell(row_idx, col_idx)
            value = str(cell.value)[:15] if cell.value else ""
            if "Articolo" in str(cell.value):
                print(f"[Col {col_idx}: {value}] *** FOUND ***", end=" ")
            else:
                print(f"[{value}]", end=" ")
        print()

    wb.close()

    # Also check source file
    source_file = Path("_ref/usbilli/Avanzamento schede 3° trimestre 2025.xlsx")
    print(f"\n\nChecking Articolo in source file: {source_file}")

    wb_src = openpyxl.load_workbook(source_file)
    ws_src = wb_src.active

    print("\nFirst row (headers):")
    for col_idx in range(1, 11):
        cell = ws_src.cell(1, col_idx)
        print(f"  Col {col_idx}: {cell.value}")

    # Show some sample Articolo and Matricola values
    print("\nSample data (Articolo vs Matricola):")
    for row_idx in range(2, 12):
        articolo = ws_src.cell(row_idx, 1).value  # Assuming col 1
        matricola = ws_src.cell(row_idx, 6).value  # Col 6
        print(f"  Row {row_idx}: Articolo={articolo}, Matricola={matricola}")

    wb_src.close()

check_articolo_in_planning()
