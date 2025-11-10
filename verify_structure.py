import openpyxl
from pathlib import Path

def verify_structure(file_path):
    """Verify the exact structure of headers and data"""
    print(f"\nVerifying structure in: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Show columns around col 31 (Rilascio DiBa/Disegni)
    print("\nColumns 29-33 (around 'Rilascio DiBa/Disegni'):")
    for row_idx in range(1, 8):
        print(f"Row {row_idx}:", end=" ")
        for col_idx in range(29, 34):
            cell = ws.cell(row_idx, col_idx)
            value = str(cell.value)[:15] if cell.value else ""
            print(f"[{col_idx}:{value}]", end=" ")
        print()

    # Show column 2 (Matricola)
    print("\nColumn 2 (Matricola) - first 10 data rows:")
    for row_idx in range(1, 11):
        cell = ws.cell(row_idx, 2)
        print(f"Row {row_idx}: {cell.value}")

    wb.close()

planning_file = Path("_ref/usbilli/Planning/Planning_25_07_04.xlsx")
verify_structure(planning_file)
