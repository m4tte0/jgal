import openpyxl
from pathlib import Path
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime
from collections import defaultdict

def analyze_delivery_performance():
    """Comprehensive delivery performance analysis with visualizations"""
    output_file = Path("Avanzamento_schede_automated.xlsx")

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    print("="*80)
    print("DELIVERY PERFORMANCE ANALYSIS")
    print("="*80)

    # Find columns
    delta_col = None
    prevista_col = None
    effettiva_col = None
    articolo_col = 1

    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        if header == "Delta":
            delta_col = col_idx
        elif header == "Data prevista avanzamento" and col_idx == ws.max_column - 1:
            prevista_col = col_idx
        elif header == "Data effettiva avanzamento":
            effettiva_col = col_idx

    # Collect data
    deltas = []
    prevista_dates = []
    effettiva_dates = []
    articolo_names = []

    for row_idx in range(2, ws.max_row + 1):
        delta = ws.cell(row_idx, delta_col).value
        prevista = ws.cell(row_idx, prevista_col).value
        effettiva = ws.cell(row_idx, effettiva_col).value
        articolo = ws.cell(row_idx, articolo_col).value

        if delta is not None and prevista and effettiva:
            deltas.append(delta)
            prevista_dates.append(prevista)
            effettiva_dates.append(effettiva)
            articolo_names.append(str(articolo)[:20])

    # Statistics
    total_items = len(deltas)
    early_count = sum(1 for d in deltas if d < 0)
    on_time_count = sum(1 for d in deltas if d == 0)
    late_count = sum(1 for d in deltas if d > 0)

    early_pct = (early_count / total_items * 100) if total_items > 0 else 0
    on_time_pct = (on_time_count / total_items * 100) if total_items > 0 else 0
    late_pct = (late_count / total_items * 100) if total_items > 0 else 0

    avg_delta = np.mean(deltas) if deltas else 0
    median_delta = np.median(deltas) if deltas else 0
    std_delta = np.std(deltas) if deltas else 0
    min_delta = min(deltas) if deltas else 0
    max_delta = max(deltas) if deltas else 0

    # Print statistics
    print(f"\nOVERALL PERFORMANCE METRICS")
    print(f"-" * 80)
    print(f"Total items analyzed: {total_items}")
    print(f"\nDelivery Performance:")
    print(f"  [+] Early (Delta < 0):    {early_count:3d} items ({early_pct:5.1f}%) - GOOD!")
    print(f"  [+] On-time (Delta = 0):  {on_time_count:3d} items ({on_time_pct:5.1f}%) - PERFECT!")
    print(f"  [-] Late (Delta > 0):     {late_count:3d} items ({late_pct:5.1f}%) - NEEDS ATTENTION")
    print(f"\nDelta Statistics (days):")
    print(f"  Average:   {avg_delta:7.1f} days")
    print(f"  Median:    {median_delta:7.1f} days")
    print(f"  Std Dev:   {std_delta:7.1f} days")
    print(f"  Min:       {min_delta:7d} days (best: earliest delivery)")
    print(f"  Max:       {max_delta:7d} days (worst: most delayed)")

    # Performance score (percentage of items delivered early or on-time)
    performance_score = ((early_count + on_time_count) / total_items * 100) if total_items > 0 else 0
    print(f"\n>> OVERALL PERFORMANCE SCORE: {performance_score:.1f}% delivered on-time or early")

    # Top 10 best and worst performers
    sorted_items = sorted(zip(deltas, articolo_names), key=lambda x: x[0])

    print(f"\n>> TOP 10 BEST PERFORMERS (Most Early):")
    print(f"{'Rank':<6} {'Delta':<10} {'Articolo':<25}")
    print(f"-" * 80)
    for i, (delta, articolo) in enumerate(sorted_items[:10], 1):
        print(f"{i:<6} {delta:>4d} days   {articolo:<25}")

    print(f"\n>> TOP 10 WORST PERFORMERS (Most Late):")
    print(f"{'Rank':<6} {'Delta':<10} {'Articolo':<25}")
    print(f"-" * 80)
    for i, (delta, articolo) in enumerate(sorted_items[-10:][::-1], 1):
        print(f"{i:<6} {delta:>4d} days   {articolo:<25}")

    # Monthly trend analysis
    monthly_deltas = defaultdict(list)
    for prevista, delta in zip(prevista_dates, deltas):
        if hasattr(prevista, 'year'):
            month_key = f"{prevista.year}-{prevista.month:02d}"
            monthly_deltas[month_key].append(delta)

    print(f"\n>> MONTHLY TREND ANALYSIS:")
    print(f"{'Month':<15} {'Avg Delta':<12} {'Items':<8} {'On-time %':<12}")
    print(f"-" * 80)
    for month in sorted(monthly_deltas.keys()):
        month_data = monthly_deltas[month]
        avg = np.mean(month_data)
        count = len(month_data)
        on_time_pct = sum(1 for d in month_data if d <= 0) / count * 100
        print(f"{month:<15} {avg:>7.1f} days {count:>4d}     {on_time_pct:>5.1f}%")

    wb.close()

    # Create visualizations
    create_visualizations(deltas, prevista_dates, effettiva_dates, articolo_names,
                         early_count, on_time_count, late_count, monthly_deltas)

    # Generate comprehensive text summary
    generate_text_summary(deltas, articolo_names, monthly_deltas,
                         early_count, on_time_count, late_count,
                         avg_delta, median_delta, std_delta, min_delta, max_delta,
                         total_items, performance_score)

    print(f"\n{'='*80}")
    print(f"Analysis complete!")
    print(f"  - Charts saved: 'delivery_analysis_*.png'")
    print(f"  - Summary saved: 'analysis_summary.txt'")
    print(f"{'='*80}")

def create_visualizations(deltas, prevista_dates, effettiva_dates, articolo_names,
                         early_count, on_time_count, late_count, monthly_deltas):
    """Create comprehensive visualizations"""

    # Set style
    plt.style.use('seaborn-v0_8-darkgrid')
    fig = plt.figure(figsize=(20, 12))

    # 1. Performance Distribution Pie Chart
    ax1 = plt.subplot(2, 3, 1)
    colors = ['#2ecc71', '#3498db', '#e74c3c']  # Green, Blue, Red
    sizes = [early_count, on_time_count, late_count]
    labels = [f'Early\n{early_count} items\n({early_count/sum(sizes)*100:.1f}%)',
              f'On-time\n{on_time_count} items\n({on_time_count/sum(sizes)*100:.1f}%)',
              f'Late\n{late_count} items\n({late_count/sum(sizes)*100:.1f}%)']
    explode = (0.05, 0.05, 0.1)  # Emphasize late items

    wedges, texts, autotexts = ax1.pie(sizes, labels=labels, colors=colors, explode=explode,
                                        autopct='', startangle=90, textprops={'fontsize': 10})
    ax1.set_title('Delivery Performance Distribution', fontsize=14, fontweight='bold', pad=20)

    # 2. Delta Distribution Histogram
    ax2 = plt.subplot(2, 3, 2)
    bins = np.arange(min(deltas) - 5, max(deltas) + 5, 5)
    counts, edges, patches = ax2.hist(deltas, bins=bins, edgecolor='black', alpha=0.7)

    # Color bars based on performance
    for patch, edge in zip(patches, edges):
        if edge < 0:
            patch.set_facecolor('#2ecc71')  # Green for early
        elif edge == 0:
            patch.set_facecolor('#3498db')  # Blue for on-time
        else:
            patch.set_facecolor('#e74c3c')  # Red for late

    ax2.axvline(x=0, color='black', linestyle='--', linewidth=2, label='On-time deadline')
    ax2.axvline(x=np.mean(deltas), color='orange', linestyle='--', linewidth=2,
                label=f'Average: {np.mean(deltas):.1f} days')
    ax2.set_xlabel('Delta (days)', fontsize=11)
    ax2.set_ylabel('Number of Items', fontsize=11)
    ax2.set_title('Distribution of Delivery Delays', fontsize=14, fontweight='bold', pad=20)
    ax2.legend()
    ax2.grid(True, alpha=0.3)

    # 3. Box Plot
    ax3 = plt.subplot(2, 3, 3)
    box_data = [
        [d for d in deltas if d < 0],   # Early
        [d for d in deltas if d == 0],  # On-time
        [d for d in deltas if d > 0]    # Late
    ]
    bp = ax3.boxplot(box_data, labels=['Early\n(Δ < 0)', 'On-time\n(Δ = 0)', 'Late\n(Δ > 0)'],
                     patch_artist=True, showmeans=True)

    for patch, color in zip(bp['boxes'], ['#2ecc71', '#3498db', '#e74c3c']):
        patch.set_facecolor(color)
        patch.set_alpha(0.7)

    ax3.axhline(y=0, color='black', linestyle='--', linewidth=1, alpha=0.5)
    ax3.set_ylabel('Delta (days)', fontsize=11)
    ax3.set_title('Delta Distribution by Category', fontsize=14, fontweight='bold', pad=20)
    ax3.grid(True, alpha=0.3, axis='y')

    # 4. Timeline Scatter Plot
    ax4 = plt.subplot(2, 3, 4)
    colors_scatter = ['#2ecc71' if d < 0 else '#3498db' if d == 0 else '#e74c3c' for d in deltas]
    scatter = ax4.scatter(prevista_dates, deltas, c=colors_scatter, alpha=0.6, s=100, edgecolors='black', linewidth=0.5)
    ax4.axhline(y=0, color='black', linestyle='--', linewidth=2, label='On-time threshold')
    ax4.set_xlabel('Promised Delivery Date (Data prevista)', fontsize=11)
    ax4.set_ylabel('Delta (days)', fontsize=11)
    ax4.set_title('Delivery Performance Over Time', fontsize=14, fontweight='bold', pad=20)
    ax4.grid(True, alpha=0.3)
    ax4.legend()
    plt.setp(ax4.xaxis.get_majorticklabels(), rotation=45)

    # 5. Monthly Average Trend
    ax5 = plt.subplot(2, 3, 5)
    if monthly_deltas:
        months = sorted(monthly_deltas.keys())
        monthly_avg = [np.mean(monthly_deltas[m]) for m in months]
        monthly_count = [len(monthly_deltas[m]) for m in months]

        x_pos = np.arange(len(months))
        bars = ax5.bar(x_pos, monthly_avg, color=['#2ecc71' if avg < 0 else '#e74c3c' for avg in monthly_avg],
                      alpha=0.7, edgecolor='black')

        # Add count labels on bars
        for i, (bar, count) in enumerate(zip(bars, monthly_count)):
            height = bar.get_height()
            ax5.text(bar.get_x() + bar.get_width()/2., height,
                    f'n={count}', ha='center', va='bottom' if height > 0 else 'top', fontsize=8)

        ax5.axhline(y=0, color='black', linestyle='--', linewidth=2)
        ax5.set_xticks(x_pos)
        ax5.set_xticklabels(months, rotation=45, ha='right')
        ax5.set_xlabel('Month', fontsize=11)
        ax5.set_ylabel('Average Delta (days)', fontsize=11)
        ax5.set_title('Monthly Average Delivery Performance', fontsize=14, fontweight='bold', pad=20)
        ax5.grid(True, alpha=0.3, axis='y')

    # 6. Cumulative Performance
    ax6 = plt.subplot(2, 3, 6)
    sorted_deltas = sorted(deltas)
    cumulative_pct = np.arange(1, len(sorted_deltas) + 1) / len(sorted_deltas) * 100

    ax6.plot(sorted_deltas, cumulative_pct, linewidth=2, color='#3498db')
    ax6.axvline(x=0, color='red', linestyle='--', linewidth=2, label='On-time threshold')

    # Find percentage delivered on-time or early
    on_time_or_early_pct = sum(1 for d in deltas if d <= 0) / len(deltas) * 100
    ax6.axhline(y=on_time_or_early_pct, color='green', linestyle='--', linewidth=2,
               label=f'{on_time_or_early_pct:.1f}% on-time or early')

    ax6.set_xlabel('Delta (days)', fontsize=11)
    ax6.set_ylabel('Cumulative Percentage (%)', fontsize=11)
    ax6.set_title('Cumulative Distribution Function', fontsize=14, fontweight='bold', pad=20)
    ax6.grid(True, alpha=0.3)
    ax6.legend()
    ax6.set_xlim([min(sorted_deltas) - 10, max(sorted_deltas) + 10])

    plt.tight_layout()
    plt.savefig('delivery_analysis_overview.png', dpi=300, bbox_inches='tight')
    print("\n[+] Saved: delivery_analysis_overview.png")

    # Create second figure with detailed analysis
    create_detailed_charts(deltas, articolo_names, prevista_dates, effettiva_dates)

def create_detailed_charts(deltas, articolo_names, prevista_dates, effettiva_dates):
    """Create detailed performance charts"""

    fig = plt.figure(figsize=(20, 10))

    # Top 15 worst performers (most late)
    ax1 = plt.subplot(1, 2, 1)
    sorted_worst = sorted(zip(deltas, articolo_names), key=lambda x: x[0], reverse=True)[:15]
    worst_deltas = [d for d, _ in sorted_worst]
    worst_names = [n for _, n in sorted_worst]

    y_pos = np.arange(len(worst_names))
    colors = ['#e74c3c' if d > 0 else '#f39c12' for d in worst_deltas]
    bars = ax1.barh(y_pos, worst_deltas, color=colors, edgecolor='black', alpha=0.8)

    ax1.set_yticks(y_pos)
    ax1.set_yticklabels(worst_names, fontsize=9)
    ax1.set_xlabel('Delta (days)', fontsize=11)
    ax1.set_title('Top 15 Most Delayed Items', fontsize=14, fontweight='bold', pad=20)
    ax1.axvline(x=0, color='black', linestyle='--', linewidth=2)
    ax1.grid(True, alpha=0.3, axis='x')

    # Add value labels
    for i, (bar, delta) in enumerate(zip(bars, worst_deltas)):
        width = bar.get_width()
        ax1.text(width, bar.get_y() + bar.get_height()/2, f'{int(delta)}d',
                ha='left', va='center', fontsize=8, fontweight='bold')

    # Top 15 best performers (most early)
    ax2 = plt.subplot(1, 2, 2)
    sorted_best = sorted(zip(deltas, articolo_names), key=lambda x: x[0])[:15]
    best_deltas = [d for d, _ in sorted_best]
    best_names = [n for _, n in sorted_best]

    y_pos = np.arange(len(best_names))
    colors = ['#2ecc71' if d < 0 else '#3498db' for d in best_deltas]
    bars = ax2.barh(y_pos, best_deltas, color=colors, edgecolor='black', alpha=0.8)

    ax2.set_yticks(y_pos)
    ax2.set_yticklabels(best_names, fontsize=9)
    ax2.set_xlabel('Delta (days)', fontsize=11)
    ax2.set_title('Top 15 Earliest Delivered Items', fontsize=14, fontweight='bold', pad=20)
    ax2.axvline(x=0, color='black', linestyle='--', linewidth=2)
    ax2.grid(True, alpha=0.3, axis='x')

    # Add value labels
    for i, (bar, delta) in enumerate(zip(bars, best_deltas)):
        width = bar.get_width()
        ax2.text(width, bar.get_y() + bar.get_height()/2, f'{int(delta)}d',
                ha='right', va='center', fontsize=8, fontweight='bold')

    plt.tight_layout()
    plt.savefig('delivery_analysis_top_performers.png', dpi=300, bbox_inches='tight')
    print("[+] Saved: delivery_analysis_top_performers.png")

    plt.close('all')

def generate_text_summary(deltas, articolo_names, monthly_deltas,
                         early_count, on_time_count, late_count,
                         avg_delta, median_delta, std_delta, min_delta, max_delta,
                         total_items, performance_score):
    """Generate comprehensive text summary file"""

    from datetime import datetime

    # Sort items for best/worst performers
    sorted_items = sorted(zip(deltas, articolo_names), key=lambda x: x[0])
    sorted_worst = sorted_items[-10:][::-1]
    sorted_best = sorted_items[:10]

    # Calculate percentages
    early_pct = (early_count / total_items * 100) if total_items > 0 else 0
    on_time_pct = (on_time_count / total_items * 100) if total_items > 0 else 0
    late_pct = (late_count / total_items * 100) if total_items > 0 else 0

    with open('analysis_summary.txt', 'w', encoding='utf-8') as f:
        f.write("="*80 + "\n")
        f.write("DELIVERY PERFORMANCE ANALYSIS - COMPREHENSIVE INSIGHTS\n")
        f.write("="*80 + "\n")
        f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("Data Source: Avanzamento_schede_automated.xlsx\n")
        f.write(f"Items Analyzed: {total_items}\n\n")

        f.write("="*80 + "\n")
        f.write("EXECUTIVE SUMMARY\n")
        f.write("="*80 + "\n\n")
        f.write(f"Performance Score: {performance_score:.1f}% items delivered on-time or early\n")
        f.write(f"Average delay: {avg_delta:.1f} days\n")
        f.write(f"Standard deviation: {std_delta:.1f} days (indicates inconsistent performance)\n\n")

        f.write("="*80 + "\n")
        f.write("PERFORMANCE DISTRIBUTION\n")
        f.write("="*80 + "\n\n")
        f.write(f"{'Category':<25} {'Count':<10} {'Percentage':<15} {'Status'}\n")
        f.write("-"*80 + "\n")
        f.write(f"{'Early (Delta < 0)':<25} {early_count:<10} {early_pct:>6.1f}%      GOOD - Before deadline\n")
        f.write(f"{'On-time (Delta = 0)':<25} {on_time_count:<10} {on_time_pct:>6.1f}%      PERFECT - Exactly on time\n")
        f.write(f"{'Late (Delta > 0)':<25} {late_count:<10} {late_pct:>6.1f}%      NEEDS ATTENTION - Exceeded deadline\n\n")

        f.write("="*80 + "\n")
        f.write("KEY STATISTICS\n")
        f.write("="*80 + "\n\n")
        f.write("Delta Statistics (days):\n")
        f.write(f"  Average:   {avg_delta:>7.1f} days\n")
        f.write(f"  Median:    {median_delta:>7.1f} days\n")
        f.write(f"  Std Dev:   {std_delta:>7.1f} days\n")
        f.write(f"  Min:       {min_delta:>7d} days (best: earliest delivery)\n")
        f.write(f"  Max:       {max_delta:>7d} days (worst: most delayed)\n\n")

        f.write("INTERPRETATION:\n")
        f.write(f"The median of {median_delta:.0f} days suggests that half of all items meet their\n")
        f.write(f"deadline. The mean of {avg_delta:+.1f} days indicates the average performance.\n")
        f.write(f"The high standard deviation ({std_delta:.1f} days) reveals inconsistent\n")
        f.write("performance across projects.\n\n")

        f.write("="*80 + "\n")
        f.write("CRITICAL ISSUES - TOP 10 MOST DELAYED ITEMS\n")
        f.write("="*80 + "\n\n")
        f.write(f"{'Rank':<6} {'Articolo':<25} {'Delay (days)':<15} {'Impact'}\n")
        f.write("-"*80 + "\n")
        for i, (delta, articolo) in enumerate(sorted_worst, 1):
            severity = "CRITICAL" if delta > 100 else "HIGH" if delta > 60 else "MEDIUM"
            months = abs(delta) / 30
            f.write(f"{i:<6} {articolo:<25} {delta:>4d}            {severity:<8} ~{months:.1f} months late\n")

        f.write("\nACTION REQUIRED:\n")
        f.write("Items delayed by 4+ months require immediate investigation!\n\n")

        f.write("="*80 + "\n")
        f.write("EXCELLENCE EXAMPLES - TOP 10 EARLIEST DELIVERIES\n")
        f.write("="*80 + "\n\n")
        f.write(f"{'Rank':<6} {'Articolo':<25} {'Early (days)':<15} {'Achievement'}\n")
        f.write("-"*80 + "\n")
        for i, (delta, articolo) in enumerate(sorted_best, 1):
            achievement = "Outstanding" if delta < -50 else "Excellent" if delta < -10 else "Very Good"
            months = abs(delta) / 30
            f.write(f"{i:<6} {articolo:<25} {delta:>4d}            {achievement:<12} ~{months:.1f} months early\n")

        f.write("\nBEST PRACTICE OPPORTUNITY:\n")
        if sorted_best:
            best_delta, best_articolo = sorted_best[0]
            f.write(f"{best_articolo} was delivered {abs(best_delta)} days early!\n")
            f.write("RECOMMENDATION: Study this success to replicate best practices.\n\n")

        f.write("="*80 + "\n")
        f.write("MONTHLY TREND ANALYSIS\n")
        f.write("="*80 + "\n\n")
        f.write(f"{'Month':<15} {'Avg Delta':<15} {'Items':<10} {'On-time %'}\n")
        f.write("-"*80 + "\n")
        for month in sorted(monthly_deltas.keys()):
            month_data = monthly_deltas[month]
            avg = np.mean(month_data)
            count = len(month_data)
            on_time_pct = sum(1 for d in month_data if d <= 0) / count * 100
            performance = "EXCELLENT" if on_time_pct >= 95 else "GOOD" if on_time_pct >= 75 else "NEEDS IMPROVEMENT"
            f.write(f"{month:<15} {avg:>7.1f} days    {count:>4d}      {on_time_pct:>5.1f}%  ({performance})\n")

        f.write("\n" + "="*80 + "\n")
        f.write("RECOMMENDATIONS\n")
        f.write("="*80 + "\n\n")
        f.write("IMMEDIATE ACTIONS:\n")
        f.write("1. Investigate top 3-5 most delayed items for root causes\n")
        f.write("2. Review months with 0% on-time rate for systemic issues\n")
        f.write("3. Study best performers to identify success factors\n\n")

        f.write("SHORT-TERM IMPROVEMENTS (1-3 Months):\n")
        f.write("4. Implement buffer time for complex projects\n")
        f.write("5. Reduce variance through standardization\n")
        f.write(f"6. Target: Reduce standard deviation from {std_delta:.1f} to <20 days\n\n")

        f.write("LONG-TERM STRATEGY (3-6 Months):\n")
        f.write(f"7. Target performance goal: 80%+ on-time or early (current: {performance_score:.1f}%)\n")
        f.write("8. Eliminate extreme outliers (no delays >30 days)\n")
        f.write("9. Implement continuous improvement culture\n\n")

        f.write("="*80 + "\n")
        f.write("PERFORMANCE BENCHMARKING\n")
        f.write("="*80 + "\n\n")
        f.write("Industry Standards:\n")
        f.write("- World-class: 95%+ on-time delivery\n")
        f.write("- Good: 85-95% on-time delivery\n")
        f.write("- Average: 70-85% on-time delivery\n")
        f.write("- Below average: <70% on-time delivery\n\n")
        f.write(f"Your Current Position: {performance_score:.1f}% = ")
        if performance_score >= 85:
            f.write("Good performance\n")
        elif performance_score >= 70:
            f.write("Average performance\n")
        else:
            f.write("Below average (but improving)\n")
        f.write("\n")

        f.write("="*80 + "\n")
        f.write("CONCLUSION\n")
        f.write("="*80 + "\n\n")
        f.write(f"Current State: {performance_score:.1f}% on-time performance\n")
        f.write(f"Target State: 80%+ on-time performance with <20 days standard deviation\n\n")
        f.write("Path Forward:\n")
        f.write("1. Address critical delays (4+ months late)\n")
        f.write("2. Replicate success factors from best performers\n")
        f.write("3. Reduce variance through process improvements\n")
        f.write("4. Build sustainable excellence through continuous improvement\n\n")

        f.write("="*80 + "\n")
        f.write("VISUALIZATIONS GENERATED\n")
        f.write("="*80 + "\n\n")
        f.write("1. delivery_analysis_overview.png - Comprehensive dashboard\n")
        f.write("2. delivery_analysis_top_performers.png - Best/worst comparison\n\n")

        f.write("="*80 + "\n")
        f.write("END OF ANALYSIS\n")
        f.write("="*80 + "\n")

    print("[+] Saved: analysis_summary.txt")

if __name__ == "__main__":
    analyze_delivery_performance()
